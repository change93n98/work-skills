#!/usr/bin/env python3
"""
Prof Analy - 模型性能分析工具
分析PyTorch profiling trace文件，生成性能分析报告
"""

import json
import gzip
import sys
import os
from collections import defaultdict
from typing import Dict, List, Tuple, Any
import argparse

# 尝试导入openpyxl，如果没有则提供安装提示
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：需要安装openpyxl库")
    print("请运行：pip install openpyxl")
    sys.exit(1)


class ProfAnalyzer:
    """性能分析器"""

    def __init__(self):
        # 算子分类规则 - 按优先级排序
        # 注意：conv_bn必须在gemm之前，因为conv算子名称中可能包含'implicitgemm'
        self.category_rules = {
            'conv_bn': ['convolution', 'batch_norm', 'conv2d', 'conv3d', 'cudnn', 'ncxdhw',
                       'conv_bn', 'implicitgemm', 'nchw2ncxhw', 'nchw2cxhwn', 'ncxhwn',
                       'nchw2ncxhw32', 'nchw2cxhwn32'],
            'attention': ['flash_fwd_kernel', 'flash_bwd_kernel', 'attention', 'scaled_dot_product',
                         'flash_attn', 'fwd_kernel', 'bwd_kernel'],
            'norm': ['layer_norm', 'rms_norm', 'group_norm', 'instance_norm', 'layernorm',
                    'rnsnorm', 'RowwiseMoments', 'NormalizeKernel', 'ComputeFusedParams',
                    'GroupNorm', 'GroupNormKernel'],
            'gemm': ['Cijk', 'gemm', 'matmul', 'bmm', 'linear', 'cublasLt', 'cublas'],
            'elementwise': ['add', 'mul', 'sub', 'div', 'relu', 'gelu', 'silu', 'sigmoid',
                          'tanh', 'softmax', 'dropout', 'where', 'clamp', 'abs', 'pow',
                          'sqrt', 'exp', 'log', 'neg', 'reciprocal', 'elementwise',
                          'vectorized_elementwise', 'FillFunctor', 'CUDAFunctor'],
            '访存': ['memcpy', 'MemCpy', 'cudaMemcpy', 'mem_set', 'memset', 'MemcpyAsync'],
            'reduction': ['sum', 'mean', 'max', 'min', 'prod', 'argmax', 'argmin',
                         'ReduceKernel', 'ReduceOp'],
            'index': ['index', 'gather', 'scatter', 'slice', 'select', 'embedding',
                     'vectorized_gather', 'indexSelectLargeIndex'],
            'shape': ['reshape', 'view', 'permute', 'transpose', 'contiguous', 'clone',
                     'CopyKernel', 'bfloat16_copy']
        }

    def load_trace(self, file_path: str) -> Dict:
        """加载trace文件"""
        print(f"加载trace文件: {file_path}")

        if file_path.endswith('.gz'):
            with gzip.open(file_path, 'rt', encoding='utf-8') as f:
                data = json.load(f)
        else:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

        print(f"加载完成，共 {len(data.get('traceEvents', []))} 个事件")
        return data

    def classify_operator(self, name: str) -> str:
        """对算子进行分类"""
        name_lower = name.lower()

        # 按优先级检查分类规则
        for category, keywords in self.category_rules.items():
            for keyword in keywords:
                if keyword.lower() in name_lower:
                    return category

        # 特殊处理一些常见的算子模式
        if 'kernel' in name_lower:
            # 尝试从kernel名称中提取更多信息
            if any(x in name_lower for x in ['gemm', 'matmul', 'bmm']):
                return 'gemm'
            elif any(x in name_lower for x in ['flash', 'attention']):
                return 'attention'
            elif any(x in name_lower for x in ['conv', 'cudnn']):
                return 'conv_bn'
            elif any(x in name_lower for x in ['norm', 'layer_norm']):
                return 'norm'
            elif any(x in name_lower for x in ['elementwise', 'ew']):
                return 'elementwise'

        return '其他'

    def analyze_events(self, events: List[Dict]) -> Tuple[Dict, Dict]:
        """分析事件数据，只关注GPU kernel事件"""
        # 算子统计
        operator_stats = defaultdict(lambda: {
            'category': '',
            'count': 0,
            'total_dur': 0.0,
            'avg_dur': 0.0
        })

        # 分类统计
        category_stats = defaultdict(lambda: {
            'operator_types': set(),
            'count': 0,
            'total_dur': 0.0
        })

        # 总耗时
        total_dur = 0.0

        # 处理每个事件 - 只关注GPU kernel事件
        for event in events:
            if event.get('ph') != 'X':  # 只处理完整事件
                continue

            name = event.get('name', '')
            dur = event.get('dur', 0.0)
            cat = event.get('cat', '')

            # 只处理kernel类别和gpu_memcpy类别的事件
            if cat not in ['kernel', 'gpu_memcpy', 'gpu_memset']:
                continue

            # 跳过一些非算子事件
            if not name or dur == 0:
                continue

            # 分类算子
            category = self.classify_operator(name)

            # 更新算子统计
            if name not in operator_stats:
                operator_stats[name]['category'] = category

            operator_stats[name]['count'] += 1
            operator_stats[name]['total_dur'] += dur

            # 更新分类统计
            category_stats[category]['operator_types'].add(name)
            category_stats[category]['count'] += 1
            category_stats[category]['total_dur'] += dur

            # 累加总耗时
            total_dur += dur

        # 计算平均耗时
        for name, stats in operator_stats.items():
            if stats['count'] > 0:
                stats['avg_dur'] = stats['total_dur'] / stats['count']

        return operator_stats, category_stats, total_dur

    def calculate_percentages(self, operator_stats: Dict, category_stats: Dict, total_dur: float):
        """计算百分比"""
        # 计算算子百分比
        for name, stats in operator_stats.items():
            stats['relative_pct'] = (stats['total_dur'] / total_dur * 100) if total_dur > 0 else 0

        # 计算分类百分比
        for category, stats in category_stats.items():
            stats['relative_pct'] = (stats['total_dur'] / total_dur * 100) if total_dur > 0 else 0

    def create_excel_report(self, operator_stats: Dict, category_stats: Dict,
                           total_dur: float, output_path: str):
        """创建Excel报告"""
        wb = Workbook()

        # ============ Sheet1: 算子详情 ============
        ws1 = wb.active
        ws1.title = "算子详情"

        # 表头
        headers1 = ['算子名称', '分类', '调用次数', '总耗时(us)', '平均耗时(us)', '相对占比(%)', '绝对占比(%)']
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 数据
        row = 2
        # 按总耗时降序排序
        sorted_operators = sorted(operator_stats.items(),
                                 key=lambda x: x[1]['total_dur'], reverse=True)

        for name, stats in sorted_operators:
            ws1.cell(row=row, column=1, value=name)
            ws1.cell(row=row, column=2, value=stats['category'])
            ws1.cell(row=row, column=3, value=stats['count'])
            ws1.cell(row=row, column=4, value=round(stats['total_dur'], 2))
            ws1.cell(row=row, column=5, value=round(stats['avg_dur'], 2))
            ws1.cell(row=row, column=6, value=round(stats['relative_pct'], 2))
            # 绝对占比（相对于总时间）
            absolute_pct = (stats['total_dur'] / total_dur * 100) if total_dur > 0 else 0
            ws1.cell(row=row, column=7, value=round(absolute_pct, 2))
            row += 1

        # 设置列宽
        ws1.column_dimensions['A'].width = 80
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 10
        ws1.column_dimensions['D'].width = 15
        ws1.column_dimensions['E'].width = 15
        ws1.column_dimensions['F'].width = 12
        ws1.column_dimensions['G'].width = 12

        # ============ Sheet2: 分类汇总 ============
        ws2 = wb.create_sheet("分类汇总")

        # 表头
        headers2 = ['分类', '算子种类数', '调用次数', '总耗时(us)', '总耗时(ms)', '相对占比(%)', '绝对占比(%)']
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 数据
        row = 2
        # 按总耗时降序排序
        sorted_categories = sorted(category_stats.items(),
                                   key=lambda x: x[1]['total_dur'], reverse=True)

        for category, stats in sorted_categories:
            ws2.cell(row=row, column=1, value=category)
            ws2.cell(row=row, column=2, value=len(stats['operator_types']))
            ws2.cell(row=row, column=3, value=stats['count'])
            ws2.cell(row=row, column=4, value=round(stats['total_dur'], 2))
            ws2.cell(row=row, column=5, value=round(stats['total_dur'] / 1000, 2))  # 转换为ms
            ws2.cell(row=row, column=6, value=round(stats['relative_pct'], 2))
            # 绝对占比
            absolute_pct = (stats['total_dur'] / total_dur * 100) if total_dur > 0 else 0
            ws2.cell(row=row, column=7, value=round(absolute_pct, 2))
            row += 1

        # 添加总计行
        row += 1
        ws2.cell(row=row, column=1, value='TOTAL')
        ws2.cell(row=row, column=1).font = Font(bold=True)

        total_operator_types = sum(len(stats['operator_types']) for stats in category_stats.values())
        total_count = sum(stats['count'] for stats in category_stats.values())

        ws2.cell(row=row, column=2, value=total_operator_types)
        ws2.cell(row=row, column=3, value=total_count)
        ws2.cell(row=row, column=4, value=round(total_dur, 2))
        ws2.cell(row=row, column=5, value=round(total_dur / 1000, 2))
        ws2.cell(row=row, column=6, value=100)
        ws2.cell(row=row, column=7, value=100)

        # 设置列宽
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 12
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['D'].width = 15
        ws2.column_dimensions['E'].width = 12
        ws2.column_dimensions['F'].width = 12
        ws2.column_dimensions['G'].width = 12

        # 保存文件
        wb.save(output_path)
        print(f"报告已保存到: {output_path}")

    def print_summary(self, operator_stats: Dict, category_stats: Dict, total_dur: float):
        """打印摘要信息"""
        print("\n" + "="*60)
        print("性能分析摘要")
        print("="*60)

        print(f"\n总耗时: {total_dur:.2f} us ({total_dur/1000:.2f} ms)")
        print(f"算子种类数: {len(operator_stats)}")
        print(f"总调用次数: {sum(s['count'] for s in operator_stats.values())}")

        print("\n分类统计 (按耗时降序):")
        print("-"*60)
        sorted_categories = sorted(category_stats.items(),
                                   key=lambda x: x[1]['total_dur'], reverse=True)

        for category, stats in sorted_categories:
            pct = (stats['total_dur'] / total_dur * 100) if total_dur > 0 else 0
            print(f"{category:15} | {len(stats['operator_types']):3}种算子 | "
                  f"{stats['count']:6}次调用 | {stats['total_dur']:12.2f} us | {pct:5.2f}%")

        print("\nTop 10 算子 (按耗时降序):")
        print("-"*60)
        sorted_operators = sorted(operator_stats.items(),
                                 key=lambda x: x[1]['total_dur'], reverse=True)[:10]

        for name, stats in sorted_operators:
            pct = (stats['total_dur'] / total_dur * 100) if total_dur > 0 else 0
            # 截断过长的算子名称
            display_name = name[:50] + "..." if len(name) > 50 else name
            print(f"{display_name:55} | {stats['count']:5}次 | {pct:5.2f}%")

    def generate_optimization_suggestions(self, category_stats: Dict, total_dur: float):
        """生成优化建议"""
        print("\n" + "="*60)
        print("优化建议")
        print("="*60)

        suggestions = []

        # 检查各个类别的占比
        for category, stats in category_stats.items():
            pct = (stats['total_dur'] / total_dur * 100) if total_dur > 0 else 0

            if category == 'gemm' and pct > 30:
                suggestions.append(f"• GEMM操作占比{pct:.1f}%较高，考虑：")
                suggestions.append("  - 使用更高效的矩阵乘法库（如cuBLASLt）")
                suggestions.append("  - 检查矩阵维度是否对齐")
                suggestions.append("  - 考虑使用混合精度训练")

            elif category == 'attention' and pct > 25:
                suggestions.append(f"• Attention操作占比{pct:.1f}%较高，考虑：")
                suggestions.append("  - 使用Flash Attention等优化实现")
                suggestions.append("  - 检查注意力头数和维度配置")
                suggestions.append("  - 考虑使用注意力优化技术")

            elif category == 'conv_bn' and pct > 20:
                suggestions.append(f"• 卷积操作占比{pct:.1f}%较高，考虑：")
                suggestions.append("  - 使用更高效的卷积算法")
                suggestions.append("  - 检查卷积核大小和步长配置")
                suggestions.append("  - 考虑使用深度可分离卷积")

            elif category == 'elementwise' and pct > 20:
                suggestions.append(f"• 逐元素操作占比{pct:.1f}%较高，考虑：")
                suggestions.append("  - 融合多个逐元素操作")
                suggestions.append("  - 使用算子融合技术")
                suggestions.append("  - 检查是否有不必要的操作")

            elif category == '访存' and pct > 10:
                suggestions.append(f"• 内存访问占比{pct:.1f}%较高，考虑：")
                suggestions.append("  - 优化数据布局")
                suggestions.append("  - 使用内存池")
                suggestions.append("  - 减少不必要的数据拷贝")

        if not suggestions:
            suggestions.append("• 性能分布较为均衡，暂无明显瓶颈")
            suggestions.append("• 建议进一步分析具体算子的实现细节")

        for suggestion in suggestions:
            print(suggestion)

    def analyze(self, input_path: str, output_path: str = None):
        """执行分析"""
        # 设置默认输出路径
        if output_path is None:
            output_path = os.path.join(os.path.dirname(input_path), 'trace_analysis.xlsx')

        # 加载trace文件
        data = self.load_trace(input_path)
        events = data.get('traceEvents', [])

        if not events:
            print("错误：没有找到trace事件")
            return

        # 分析事件
        print("分析算子数据...")
        operator_stats, category_stats, total_dur = self.analyze_events(events)

        # 计算百分比
        self.calculate_percentages(operator_stats, category_stats, total_dur)

        # 打印摘要
        self.print_summary(operator_stats, category_stats, total_dur)

        # 生成优化建议
        self.generate_optimization_suggestions(category_stats, total_dur)

        # 创建Excel报告
        print("\n生成Excel报告...")
        self.create_excel_report(operator_stats, category_stats, total_dur, output_path)

        print("\n分析完成！")


def main():
    parser = argparse.ArgumentParser(description='模型性能分析工具')
    parser.add_argument('input', help='trace.json或trace.json.gz文件路径')
    parser.add_argument('-o', '--output', help='输出Excel文件路径（默认为trace_analysis.xlsx）')

    args = parser.parse_args()

    # 检查输入文件是否存在
    if not os.path.exists(args.input):
        print(f"错误：找不到输入文件 {args.input}")
        sys.exit(1)

    # 执行分析
    analyzer = ProfAnalyzer()
    analyzer.analyze(args.input, args.output)


if __name__ == '__main__':
    main()