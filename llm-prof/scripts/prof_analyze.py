#!/usr/bin/env python3
"""
Profiling trace analyzer for vLLM/SGLang inference on DCU/ROCm.
Analyzes Chrome trace JSON files to categorize operator timing by phase (prefill/decode).

Usage:
    python3 prof_analyze.py --trace-file <path> [--output-dir <dir>] [--decode-step 2] [--verbose]
"""

import argparse
import gzip
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Operator classification patterns (case-insensitive matching)
OP_CATEGORIES = {
    "gemm": [
        "gemm", "Gemm", "GEMM", "hgemm",
        "Hipblaslt_Launch", "hipblaslt", "rocblas", "ROCBLAS",
        "CKGemm", "ck_gemm", "DeviceGemm", "GroupedGemm",
        "SplitkGemm", "StreamkGemm",
        "Cijk", "cijk",  # rocBLAS kernel naming
    ],
    "通信 (comm)": [
        "AllReduce", "all_reduce", "nccl", "rccl", "NCCL", "RCCL",
        "broadcast", "Broadcast", "AllGather", "all_gather",
        "ReduceScatter", "reduce_scatter", "CustomAllReduce",
        "allreduce", "AllReduceRing", "AllReduceTree",
    ],
    "FlashAttention (fa)": [
        "flash_attn", "flash_fwd", "flash_bwd", "fmha", "fmoe",
        "FlashAttention", "flash_attention", "FlashDecoding",
        "flash_decoding", "MHA", "mha_fwd", "mha_bwd",
        "ck_fmha", "CKFmha",
    ],
    "Triton": [
        "triton", "Triton", "triton_kernel",
    ],
    "其他elementwise": [
        "elementwise", "ElementWise",
        "vectorized",  # PyTorch vectorized kernels
        "softmax", "Softmax", "layernorm", "LayerNorm",
        "rmsnorm", "RMSNorm", "SiluGelu", "silu", "gelu", "relu",
        "embedding", "Embedding", "rope", "RoPE", "RotaryEmbedding",
        "topk", "TopK", "top_k", "sampling", "Sampling",
        "transpose", "reshape", "view", "contiguous",
        "copy_", "clone", "fill_", "scale", "Scale",
        "reduce_kernel",  # PyTorch reduce kernels (not AllReduce/ReduceScatter)
        "unary", "binary", "ternary",
    ],
    "memcpy/memset": [
        "memcpy", "memset", "Memcpy", "Memset", "MemCpy", "MemSet",
        "D2H", "H2D", "H2H", "D2D",
        "hipMemcpy", "hipMemset", "hipMem",
        "AsyncMemcpy", "async_memcpy", "MemcpyAsync", "memcpy_async",
    ],
}

EXCLUDE_PATTERNS = [
    "profiler_step", "ProfilerStep", "profiler",
    "Profile", "PROFILER", "torch.autograd", "autograd",
]


def load_trace(trace_path: str) -> list:
    """Load trace file and return list of events."""
    path = Path(trace_path)
    if path.suffix == ".gz":
        with gzip.open(path, "rt", encoding="utf-8") as f:
            data = json.load(f)
    else:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)

    if isinstance(data, list):
        return data
    elif isinstance(data, dict):
        if "traceEvents" in data:
            return data["traceEvents"]
        elif "events" in data:
            return data["events"]
        return [data]
    return []


def find_model_forward_slices(events: list) -> list:
    """Find all model_forward slices to identify prefill/decode boundaries."""
    forward_slices = []

    for event in events:
        if not isinstance(event, dict):
            continue
        name = str(event.get("name", ""))
        if "model_forward" in name or "ModelForward" in name:
            ts = event.get("ts", 0)
            dur = event.get("dur", 0)
            if dur > 0:
                forward_slices.append({
                    "name": name,
                    "ts": ts,
                    "dur": dur,
                    "end": ts + dur,
                })

    forward_slices.sort(key=lambda x: x["ts"])
    return forward_slices


def find_phases(events: list, decode_step: int = 2) -> dict:
    """Identify prefill and decode phase time boundaries.

    Chrome trace timestamps are in microseconds.
    Detects slow vs fast decode steps by finding the last large gap (>100ms)
    between consecutive model_forward slices. Steps after this gap are the
    fast decode cadence.

    Prefill = trace_start → first fast step start (includes warmup gap)
    Decode = the Nth fast step (1-indexed: decode_step=2 → 2nd fast step)
    """
    forward_slices = find_model_forward_slices(events)

    if not forward_slices:
        print("WARNING: No model_forward slices found")
        return {}

    trace_start = min(e.get("ts", float("inf")) for e in events
                      if isinstance(e, dict) and "ts" in e)

    # Find the last large gap (>100ms) between consecutive slices.
    # This separates slow/warmup steps from fast decode steps.
    transition_idx = -1
    for i in range(len(forward_slices) - 1):
        gap = forward_slices[i + 1]["ts"] - forward_slices[i]["end"]
        if gap > 100_000:  # > 100ms in microseconds
            transition_idx = i

    if transition_idx >= 0:
        # Prefill: trace start → first fast step start (includes gap)
        prefill_start = trace_start
        prefill_end = forward_slices[transition_idx + 1]["ts"]

        # Fast decode steps: everything after the last large gap
        fast_steps = forward_slices[transition_idx + 1:]

        # decode_step N: fast_steps[N].start → fast_steps[N+1].start
        # (1-indexed: decode_step=2 means the 2nd fast step)
        if decode_step < len(fast_steps) - 1:
            decode_start = fast_steps[decode_step]["ts"]
            decode_end = fast_steps[decode_step + 1]["ts"]
        elif fast_steps:
            print(f"WARNING: decode_step={decode_step} out of range "
                  f"({len(fast_steps)} fast steps), using last")
            decode_start = fast_steps[-1]["ts"]
            decode_end = fast_steps[-1]["end"]
        else:
            print("ERROR: No fast steps found after transition")
            return {}

        return {
            "prefill": {"start": prefill_start, "end": prefill_end},
            "decode": {"start": decode_start, "end": decode_end},
            "forward_slices": forward_slices,
            "transition_idx": transition_idx,
            "fast_steps_count": len(fast_steps),
        }
    else:
        # No large gap found → original simple logic
        prefill_start = forward_slices[0]["ts"]
        prefill_end = forward_slices[0]["end"]

        if len(forward_slices) > decode_step:
            decode_start = forward_slices[decode_step]["ts"]
            decode_end = forward_slices[decode_step]["end"]
        elif len(forward_slices) > 1:
            print(f"WARNING: Only {len(forward_slices)} forward slices, using last as decode")
            decode_start = forward_slices[-1]["ts"]
            decode_end = forward_slices[-1]["end"]
        else:
            print("ERROR: Only 1 forward slice, cannot separate prefill/decode")
            return {}

        return {
            "prefill": {"start": prefill_start, "end": prefill_end},
            "decode": {"start": decode_start, "end": decode_end},
            "forward_slices": forward_slices,
        }


def get_gpu_kernels(events: list, start_ts: int, end_ts: int) -> list:
    """Extract GPU kernel events within a time range.

    In Chrome trace format, GPU kernels have cat="kernel".
    CPU ops that launch GPU kernels have cat="cpu_op".
    """
    kernels = []

    for event in events:
        if not isinstance(event, dict):
            continue

        name = str(event.get("name", ""))
        cat = str(event.get("cat", ""))
        ts = event.get("ts", 0)
        dur = event.get("dur", 0)

        # GPU kernels: cat is "kernel" (actual GPU execution)
        # Also include "gpu_memcpy" and "gpu_memset"
        is_gpu_kernel = cat in ("kernel", "gpu_memcpy", "gpu_memset")

        # Also check for CPU ops that represent GPU kernel launches
        # (these have the kernel name but are CPU-side timing)
        is_cpu_gpu_op = cat == "cpu_op" and any(
            kw in name for kw in [
                "flash", "Flash", "gemm", "Gemm", "GEMM",
                "AllReduce", "all_reduce", "nccl", "rccl",
                "triton", "Triton", "ck_", "CK_",
                "fmha", "FMHA", "softmax", "Softmax",
                "layernorm", "LayerNorm", "rmsnorm", "RMSNorm",
                "silu", "gelu", "relu", "rope", "RoPE",
                "topk", "TopK", "embedding", "Embedding",
                "memcpy", "memset", "Memcpy", "Memset",
            ]
        )

        if not (is_gpu_kernel or is_cpu_gpu_op):
            continue

        # Filter by time range
        if ts < start_ts or (ts + dur) > end_ts:
            continue

        if dur <= 0:
            continue

        # Exclude profiling overhead
        if any(p.lower() in name.lower() for p in EXCLUDE_PATTERNS):
            continue

        kernels.append({
            "name": name,
            "ts": ts,
            "dur": dur,  # in microseconds
            "cat": cat,
        })

    return kernels


def classify_kernel(name: str) -> str:
    """Classify a kernel name into one of the 6 categories."""
    name_lower = name.lower()

    # Check exclusions first
    for pattern in EXCLUDE_PATTERNS:
        if pattern.lower() in name_lower:
            return None

    for category, patterns in OP_CATEGORIES.items():
        for pattern in patterns:
            if pattern.lower() in name_lower:
                return category

    return "其他elementwise"


def analyze_kernels(kernels: list) -> dict:
    """Analyze kernels and compute statistics per category.

    Duration is in microseconds (Chrome trace standard).
    """
    stats = defaultdict(lambda: {"total_dur": 0, "count": 0})

    for kernel in kernels:
        category = classify_kernel(kernel["name"])
        if category is None:
            continue
        stats[category]["total_dur"] += kernel["dur"]
        stats[category]["count"] += 1

    total_dur = sum(s["total_dur"] for s in stats.values())
    result = {}

    for cat in list(OP_CATEGORIES.keys()) + ["总计"]:
        if cat in stats:
            s = stats[cat]
            result[cat] = {
                "total_dur_ms": s["total_dur"] / 1000,  # us -> ms
                "count": s["count"],
                "percentage": (s["total_dur"] / total_dur * 100) if total_dur > 0 else 0,
            }
        elif cat == "总计":
            result[cat] = {
                "total_dur_ms": total_dur / 1000,
                "count": sum(s["count"] for s in stats.values()),
                "percentage": 100.0,
            }
        else:
            result[cat] = {"total_dur_ms": 0, "count": 0, "percentage": 0}

    return result


def format_table(phase_name: str, stats: dict) -> str:
    """Format analysis results as a table."""
    lines = []
    lines.append(f"\n{'=' * 60}")
    lines.append(f"  {phase_name}算子耗时分析")
    lines.append(f"{'=' * 60}")
    lines.append(f"{'类别':<20s} {'总耗时(ms)':>12s} {'调用次数':>10s} {'占比(%)':>10s}")
    lines.append(f"{'-' * 60}")

    order = ["gemm", "通信 (comm)", "FlashAttention (fa)", "Triton",
             "其他elementwise", "memcpy/memset", "总计"]

    for cat in order:
        if cat in stats:
            s = stats[cat]
            if cat == "总计":
                lines.append(f"{'-' * 60}")
            lines.append(
                f"{cat:<20s} {s['total_dur_ms']:>12.2f} {s['count']:>10d} {s['percentage']:>9.2f}%"
            )

    lines.append(f"{'=' * 60}")
    return "\n".join(lines)


def analyze_kernels_detailed(kernels: list, phase_duration_us: float) -> list:
    """Analyze kernels with detailed per-kernel statistics.

    Returns list of dicts with:
    - kernel_name, category, call_count, total_duration_us, avg_duration_us
    - relative_pct: percentage of total kernel time
    - absolute_pct: percentage of phase window time (including gaps)
    """
    kernel_stats = defaultdict(lambda: {"count": 0, "total_dur": 0})

    for kernel in kernels:
        category = classify_kernel(kernel["name"])
        if category is None:
            continue
        key = kernel["name"]
        kernel_stats[key]["count"] += 1
        kernel_stats[key]["total_dur"] += kernel["dur"]
        kernel_stats[key]["category"] = category

    total_kernel_dur = sum(s["total_dur"] for s in kernel_stats.values())

    result = []
    for name, stats in kernel_stats.items():
        result.append({
            "kernel_name": name,
            "category": stats["category"],
            "call_count": stats["count"],
            "total_duration_us": stats["total_dur"],
            "avg_duration_us": stats["total_dur"] / stats["count"] if stats["count"] > 0 else 0,
            "relative_pct": (stats["total_dur"] / total_kernel_dur * 100) if total_kernel_dur > 0 else 0,
            "absolute_pct": (stats["total_dur"] / phase_duration_us * 100) if phase_duration_us > 0 else 0,
        })

    result.sort(key=lambda x: x["total_duration_us"], reverse=True)
    return result


def save_xlsx(output_dir: str, prefill_detailed: list, decode_detailed: list,
              prefill_stats: dict, decode_stats: dict, decode_step: int):
    """Save results to xlsx with 4 sheets."""
    if not HAS_OPENPYXL:
        print("WARNING: openpyxl not installed, skipping xlsx output")
        print("  Install with: pip install openpyxl")
        return None

    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_align = Alignment(horizontal="right")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Category colors for summary sheets
    cat_colors = {
        "gemm": "FFD700",
        "通信 (comm)": "87CEEB",
        "FlashAttention (fa)": "98FB98",
        "Triton": "DDA0DD",
        "其他elementwise": "F0E68C",
        "memcpy/memset": "FFA07A",
        "总计": "C0C0C0",
    }

    def write_detail_sheet(ws, title, data):
        """Write a detailed kernel sheet."""
        ws.title = title

        headers = ["算子名称", "分类", "调用次数", "总耗时(us)", "平均耗时(us)",
                    "相对占比(%)", "绝对占比(%)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for row, item in enumerate(data, 2):
            ws.cell(row=row, column=1, value=item["kernel_name"]).border = thin_border
            ws.cell(row=row, column=2, value=item["category"]).border = thin_border
            ws.cell(row=row, column=3, value=item["call_count"]).border = thin_border
            ws.cell(row=row, column=3).alignment = num_align

            c4 = ws.cell(row=row, column=4, value=round(item["total_duration_us"], 2))
            c4.border = thin_border
            c4.number_format = '#,##0.00'

            c5 = ws.cell(row=row, column=5, value=round(item["avg_duration_us"], 2))
            c5.border = thin_border
            c5.number_format = '#,##0.00'

            c6 = ws.cell(row=row, column=6, value=round(item["relative_pct"], 2))
            c6.border = thin_border
            c6.number_format = '0.00'

            c7 = ws.cell(row=row, column=7, value=round(item["absolute_pct"], 2))
            c7.border = thin_border
            c7.number_format = '0.00'

        # Auto width
        for col in range(1, len(headers) + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(data) + 2))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 60)

    def write_summary_sheet(ws, title, detailed_data, phase_stats):
        """Write a summary sheet grouped by category."""
        ws.title = title

        headers = ["分类", "算子种类数", "调用次数", "总耗时(us)", "总耗时(ms)", "相对占比(%)"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Aggregate by category
        cat_agg = defaultdict(lambda: {"kernel_count": 0, "call_count": 0, "total_dur": 0})
        for item in detailed_data:
            cat = item["category"]
            cat_agg[cat]["kernel_count"] += 1
            cat_agg[cat]["call_count"] += item["call_count"]
            cat_agg[cat]["total_dur"] += item["total_duration_us"]

        total_dur = sum(v["total_dur"] for v in cat_agg.values())

        order = ["gemm", "通信 (comm)", "FlashAttention (fa)", "Triton",
                 "其他elementwise", "memcpy/memset"]
        row = 2
        for cat in order:
            if cat in cat_agg:
                agg = cat_agg[cat]
                pct = (agg["total_dur"] / total_dur * 100) if total_dur > 0 else 0

                ws.cell(row=row, column=1, value=cat).border = thin_border
                ws.cell(row=row, column=2, value=agg["kernel_count"]).border = thin_border
                ws.cell(row=row, column=2).alignment = num_align
                ws.cell(row=row, column=3, value=agg["call_count"]).border = thin_border
                ws.cell(row=row, column=3).alignment = num_align

                c4 = ws.cell(row=row, column=4, value=round(agg["total_dur"], 2))
                c4.border = thin_border
                c4.number_format = '#,##0.00'

                c5 = ws.cell(row=row, column=5, value=round(agg["total_dur"] / 1000, 2))
                c5.border = thin_border
                c5.number_format = '#,##0.00'

                c6 = ws.cell(row=row, column=6, value=round(pct, 2))
                c6.border = thin_border
                c6.number_format = '0.00'

                # Color the category cell
                if cat in cat_colors:
                    ws.cell(row=row, column=1).fill = PatternFill(
                        start_color=cat_colors[cat], end_color=cat_colors[cat], fill_type="solid"
                    )

                row += 1

        # Total row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).font = Font(bold=True)

        ws.cell(row=row, column=1, value="总计")
        ws.cell(row=row, column=2, value=sum(v["kernel_count"] for v in cat_agg.values()))
        ws.cell(row=row, column=3, value=sum(v["call_count"] for v in cat_agg.values()))
        ws.cell(row=row, column=4, value=round(total_dur, 2)).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=round(total_dur / 1000, 2)).number_format = '#,##0.00'
        ws.cell(row=row, column=6, value=100.00).number_format = '0.00'

        # Auto width
        for col in range(1, len(headers) + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    # Sheet 1: Prefill detailed
    ws1 = wb.active
    write_detail_sheet(ws1, "Prefill详细算子", prefill_detailed)

    # Sheet 2: Prefill summary
    ws2 = wb.create_sheet()
    write_summary_sheet(ws2, "Prefill分类汇总", prefill_detailed, prefill_stats)

    # Sheet 3: Decode detailed
    ws3 = wb.create_sheet()
    write_detail_sheet(ws3, f"Decode-Step{decode_step}详细算子", decode_detailed)

    # Sheet 4: Decode summary
    ws4 = wb.create_sheet()
    write_summary_sheet(ws4, f"Decode-Step{decode_step}分类汇总", decode_detailed, decode_stats)

    xlsx_path = os.path.join(output_dir, "prof_analysis.xlsx")
    wb.save(xlsx_path)
    return xlsx_path


def save_results(output_dir: str, prefill_stats: dict, decode_stats: dict,
                 prefill_kernels: list, decode_kernels: list, phases: dict,
                 decode_step: int):
    """Save analysis results to files."""
    os.makedirs(output_dir, exist_ok=True)

    summary_path = os.path.join(output_dir, "prof_analysis_summary.txt")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("大模型推理Profiling算子耗时分析报告\n")
        f.write(f"生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        if phases and "forward_slices" in phases:
            slices = phases["forward_slices"]
            trans = phases.get("transition_idx", -1)
            fast_count = phases.get("fast_steps_count", 0)
            f.write(f"\n检测到 {len(slices)} 个forward step:\n")
            if trans >= 0:
                f.write(f"  慢 steps: 0-{trans} (共 {trans+1} 个)\n")
                f.write(f"  快 steps: {trans+1}-{len(slices)-1} (共 {fast_count} 个)\n")
                f.write(f"  分界 gap: {(slices[trans+1]['ts'] - slices[trans]['end'])/1000:.2f} ms\n\n")
            for i, fs in enumerate(slices):
                if trans >= 0:
                    if i <= trans:
                        label = f"slow-{i}"
                    else:
                        label = f"fast-{i - trans - 1}"
                else:
                    label = "prefill" if i == 0 else f"decode-{i}"
                gap_ms = (slices[i]["ts"] - slices[i-1]["end"]) / 1000 if i > 0 else 0
                gap_str = f" gap={gap_ms:.1f}ms" if i > 0 and gap_ms > 50 else ""
                f.write(f"  Step {i} ({label}): dur={fs['dur']/1000:.2f} ms{gap_str}\n")

        f.write(format_table("Prefill阶段", prefill_stats))
        f.write(format_table(f"Decode阶段 (Step {decode_step})", decode_stats))
        f.write("\n")

    json_path = os.path.join(output_dir, "prof_analysis.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "prefill": prefill_stats,
            "decode": decode_stats,
            "prefill_kernel_count": len(prefill_kernels),
            "decode_kernel_count": len(decode_kernels),
        }, f, indent=2, ensure_ascii=False)

    return summary_path, json_path


def main():
    parser = argparse.ArgumentParser(description="Analyze vLLM/SGLang profiling traces")
    parser.add_argument("--trace-file", required=True, help="Path to trace JSON(.gz)")
    parser.add_argument("--output-dir", default=None, help="Output directory")
    parser.add_argument("--decode-step", type=int, default=2,
                        help="Which decode step to analyze (default: 2)")
    parser.add_argument("--verbose", action="store_true")

    args = parser.parse_args()

    if args.output_dir is None:
        args.output_dir = os.path.dirname(os.path.abspath(args.trace_file))

    print(f"Loading trace: {args.trace_file}")
    events = load_trace(args.trace_file)
    print(f"Loaded {len(events)} events")

    # Find phases
    print("Identifying prefill/decode phases...")
    phases = find_phases(events, decode_step=args.decode_step)

    if not phases or "prefill" not in phases:
        print("ERROR: Could not identify phases in trace")
        sys.exit(1)

    prefill = phases["prefill"]
    decode = phases["decode"]

    print(f"Prefill: {(prefill['end'] - prefill['start']) / 1000:.2f} ms")
    print(f"Decode (step {args.decode_step}): {(decode['end'] - decode['start']) / 1000:.2f} ms")

    if phases.get("forward_slices"):
        slices = phases["forward_slices"]
        trans = phases.get("transition_idx", -1)
        fast_count = phases.get("fast_steps_count", 0)
        print(f"Total forward steps: {len(slices)}")
        if trans >= 0:
            print(f"  Slow steps: 0-{trans} ({trans+1} steps)")
            print(f"  Fast steps: {trans+1}-{len(slices)-1} ({fast_count} steps)")
            gap = (slices[trans+1]["ts"] - slices[trans]["end"]) / 1000
            print(f"  Transition gap: {gap:.2f} ms")
        for i, fs in enumerate(slices):
            if trans >= 0:
                label = f"slow-{i}" if i <= trans else f"fast-{i - trans - 1}"
            else:
                label = f"step-{i}"
            gap_ms = (slices[i]["ts"] - slices[i-1]["end"]) / 1000 if i > 0 else 0
            gap_str = f" gap={gap_ms:.1f}ms" if i > 0 and gap_ms > 50 else ""
            print(f"  Step {i} ({label}): dur={fs['dur']/1000:.2f} ms{gap_str}")

    # Extract kernels
    print("\nExtracting GPU kernels...")
    prefill_kernels = get_gpu_kernels(events, prefill["start"], prefill["end"])
    decode_kernels = get_gpu_kernels(events, decode["start"], decode["end"])

    print(f"Prefill: {len(prefill_kernels)} kernels, Decode: {len(decode_kernels)} kernels")

    if args.verbose:
        # Show category distribution by count
        from collections import Counter
        prefill_cats = Counter(classify_kernel(k["name"]) for k in prefill_kernels)
        decode_cats = Counter(classify_kernel(k["name"]) for k in decode_kernels)
        print(f"\nPrefill categories: {dict(prefill_cats)}")
        print(f"Decode categories: {dict(decode_cats)}")

        print("\nTop 5 prefill kernels by duration:")
        sorted_prefill = sorted(prefill_kernels, key=lambda x: x["dur"], reverse=True)
        for k in sorted_prefill[:5]:
            print(f"  {k['name'][:60]:60s} {k['dur']/1000:>8.2f} ms  [{classify_kernel(k['name'])}]")

    # Analyze
    print("\nAnalyzing operator categories...")
    prefill_stats = analyze_kernels(prefill_kernels)
    decode_stats = analyze_kernels(decode_kernels)

    # Detailed analysis per kernel
    prefill_duration = prefill["end"] - prefill["start"]
    decode_duration = decode["end"] - decode["start"]
    prefill_detailed = analyze_kernels_detailed(prefill_kernels, prefill_duration)
    decode_detailed = analyze_kernels_detailed(decode_kernels, decode_duration)

    # Output
    print(format_table("Prefill阶段", prefill_stats))
    print(format_table(f"Decode阶段 (Step {args.decode_step})", decode_stats))

    # Save
    summary_path, json_path = save_results(
        args.output_dir, prefill_stats, decode_stats,
        prefill_kernels, decode_kernels, phases, args.decode_step
    )
    print(f"\nResults saved to:")
    print(f"  Summary: {summary_path}")
    print(f"  JSON: {json_path}")

    # Save xlsx
    xlsx_path = save_xlsx(
        args.output_dir, prefill_detailed, decode_detailed,
        prefill_stats, decode_stats, args.decode_step
    )
    if xlsx_path:
        print(f"  XLSX: {xlsx_path}")


if __name__ == "__main__":
    main()
